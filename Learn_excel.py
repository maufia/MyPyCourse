"""Learn read/wrtire to excell spreadsheet"""

import pprint
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

pp = pprint.PrettyPrinter(indent=4)

today_date = date.today()


def my_function() -> dict:
    """Return a set of data into a dictionary"""
    hills_of_rome = {
        'Aventine Hill': {'Latin': 'Aventinus',
                          'Height': 46.6,
                          'Italian': 'Aventino'},
        'Caelian Hill': {'Latin': r'Cælius',
                         'Height': 50.0,
                         'Italian': 'Celio'},
        'Capitoline Hill': {'Latin': 'Capitolinus',
                            'Height': 44.7,
                            'Italian': 'Campidoglio'},
        'Esquiline Hill': {'Latin': 'Esquilinus',
                           'Height': 58.3,
                           'Italian': 'Esquilino'},
        'Palatine Hill': {'Latin': 'Palatinus',
                          'Height': 51.0,
                          'Italian': 'Palatino'},
        'Quirinal Hill': {'Latin': 'Quirinalis',
                          'Height': 50.9,
                          'Italian': 'Quirinale'},
        'Viminal Hill': {'Latin': 'Viminalis',
                         'Height': 57.0,
                         'Italian': 'Viminale'}}
    return hills_of_rome


def set_cell_styles() -> dict:
    """
    Options for cells

    :return: dictionary with various cell style options
    """
    # Set fonts for title and text
    title_font = Font(name='Calibri', size=14, bold=True, italic=False,
                      vertAlign=None, underline='single', strike=False,
                      color='00FF00')
    text_font = Font(name='Calibri', size=12, bold=False, italic=False,
                     vertAlign=None, underline=None, strike=False,
                     color='000000')
    # Set alignments for title and text
    title_alignment = Alignment(horizontal='center', vertical='center',
                                wrapText=False)
    text_alignment = Alignment(horizontal='left', vertical='center',
                               wrapText=False)
    border_size = Side(style='thin', color='000000')
    borders_size = Border(left=border_size, right=border_size,
                          top=border_size, bottom=border_size)
    fill_type = PatternFill("solid", fgColor="DDDDDD")
    return {'title_font': title_font,
            'text_font': text_font,
            'title_alignment': title_alignment,
            'text_alignment': text_alignment,
            'borders_size': borders_size,
            'fill_type': fill_type}


def write_history_sheet(hist_sheet: dict) -> True:
    """
    Write history sheet
    :param hist_sheet: workbook'sheet
    :return:
    """
    cell_styles: dict = set_cell_styles()

    # Write to cell
    hist_sheet['A1'] = 'Date'
    hist_sheet['B1'] = 'Author'
    hist_sheet['C1'] = 'Location'
    # set cell styles for a title cell
    for cell in ["A1", "B1", "C1"]:
        hist_sheet[cell].font = cell_styles['title_font']
        hist_sheet[cell].alignment = cell_styles['title_alignment']
        hist_sheet[cell].fill = cell_styles['fill_type']
        hist_sheet[cell].border = cell_styles['borders_size']

    hist_sheet['A2'] = today_date.strftime("%d/%m/%Y")
    hist_sheet['B2'] = 'Julius Caesar'
    hist_sheet['C2'] = 'Rome'
    hist_sheet['A3'] = '01/12/2018'
    hist_sheet['B3'] = 'Marcus Aurelio'
    hist_sheet['C3'] = 'Rome'

    # set cell styles for a normal text cell
    for col in ["A", "B", "C"]:
        for row in range(2, 4):
            # Create a cell as "A2", "B2" etc
            cell = col + str(row)
            # assign properties of a cell
            hist_sheet[cell].font = cell_styles['text_font']
            hist_sheet[cell].alignment = cell_styles['text_alignment']
            hist_sheet[cell].border = cell_styles['borders_size']
            if col == 'A':
                hist_sheet[cell].font = Font(bold=True)
    return True


def write_hills_sheet(hills_sheet: load_workbook, hills_of_rome: dict) -> True:
    """
    Write hills sheet
    :param hills_of_rome: dictionary with details of hills of Rome
    :param hills_sheet: workbook sheet for hills of Rome
    :return: True
    """
    # Create the first row with headings
    headings = ['English Name', 'Italian Name', 'Latin Name', 'Height [m]']
    # for each pattern name write to a column
    for column, heading in enumerate(headings):
        column = column + 1  # excell starts counting from 1
        # Write the name of the patten in row 1
        hills_sheet.cell(row=1, column=column).value = heading
        hills_sheet.cell(row=1, column=column).font = Font(bold=True)
        hills_sheet.cell(row=1, column=column).fill = PatternFill(
            "solid", fgColor="DDDDDD")

    # for each domains discovered in the search (also sorted)
    for row_count, hill in enumerate(sorted(hills_of_rome)):
        # write the hill in subsequent rows (same column as above)
        hills_sheet.cell(row=row_count + 2, column=1).value = hill
        hills_sheet.cell(row=row_count + 2, column=2).value = \
            hills_of_rome[hill]['Italian']
        hills_sheet.cell(row=row_count + 2, column=3).value = \
            hills_of_rome[hill]['Latin']
        hills_sheet.cell(row=row_count + 2, column=4).value = \
            hills_of_rome[hill]['Height']
    return True


def set_workbook_properties(workbook: Workbook) -> True:
    """
    Set properties of excel workbook
    :param workbook: Workbook
    :return:
    """
    # To get old properties
    # obj = workbook.properties
    # print(obj)

    workbook.properties.creator = "Virgil"
    workbook.properties.title = "Hills Of Rome"
    workbook.properties.description = "Names of the Hills of Rome"
    workbook.properties.subject = "Hills and Names"
    workbook.properties.category = "Geographical"
    workbook.properties.subject = "newSubject"
    workbook.properties.created = datetime.now()
    workbook.properties.keywords = "Hills, Height, Rome"
    return True


def write_to_spreadsheet(filename: str, hills_of_rome: dict) -> True:
    """
    Write the dictionary to spreadsheet

    :param filename: string filename
    :param hills_of_rome: dictionary with hills of rome

    :return: True
    """
    # open workbook in memory
    workbook = Workbook()
    # Creates history sheet named 'History'
    hist_sheet = workbook.create_sheet(title='History')
    write_history_sheet(hist_sheet)
    # Creates sheet named 'Hills'
    hills_sheet = workbook.create_sheet(title='Hills')
    write_hills_sheet(hills_sheet, hills_of_rome)
    # Set workbook properties
    set_workbook_properties(workbook)
    # delete default Sheet
    del workbook["Sheet"]
    print("Writing to file")
    # Write workbook to file
    workbook.save(filename)
    return True


def read_spreadsheet(filename: str, sheetname: str) -> dict:
    """
    Read the dictionary to spreadsheet

    :param filename: string filename
    :param sheetname: name of the sheet to read

    return: hills of rome dictionary
    """
    print('Reading spreadsheet.')
    # Prepare dictionary
    hills_of_rome = {}
    try:
        # Open notebook
        workbook = load_workbook(filename=filename)
        # Open sheet with data
        hills_sheet: Workbook = workbook[sheetname]
    except FileNotFoundError:
        print(f'File not found: {filename}')
    except KeyError:
        print(f'Sheet "{sheetname}" not found')
    finally:
        # read every row
        for row_count, sheet_row in enumerate(hills_sheet.rows):
            if row_count == 0:
                continue
            hills_of_rome[sheet_row[0].value] = {
                'Italian': sheet_row[1].value,
                'Latin': sheet_row[2].value,
                'Height': sheet_row[3].value}
    return hills_of_rome


def main():
    filename = 'hills_of_rome.xlsx'
    print('Read data.')
    hills_of_rome = my_function()
    # Write to spreadsheet
    write_to_spreadsheet(filename, hills_of_rome)
    # Read back the spreadsheet
    hor = read_spreadsheet(filename=filename, sheetname='Hills')
    print('Print spreadsheet content:')
    pp.pprint(hor)


# --------------------------------------------------


if __name__ == "__main__":
    main()
